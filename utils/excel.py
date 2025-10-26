from openpyxl import load_workbook
import pandas as pd

def to_wb_and_ws(path, sheet_name, size=False):
    if size:
        pos1 = size[0]
        pos2 = size[1]
        wb = load_workbook(path)
        ws = wb[sheet_name]
        return wb, ws[pos1:pos2]
    else:
        wb = load_workbook(path)
        ws = wb[sheet_name]
        return wb, ws
def to_wb(path):
    wb = load_workbook(path)
    return wb
def to_ws(path, sheet_name, size=False):
    if size:
        pos1 = size[0]
        pos2 = size[1]
        wb = load_workbook(path)
        ws = wb[sheet_name]
        return ws[pos1:pos2]
    else:
        wb = load_workbook(path)
        ws = wb[sheet_name]
        return ws

def ws_to_df(ws, dataonly=False):
    if not dataonly:
        return pd.DataFrame(list(ws))

    # Barcha satrlarni qiymatlari bilan olish
    data = list(ws)

    # Birinchi qator — sarlavhalar
    columns = data[0]
    rows = data[1:]

    # DataFrame yaratish
    df = pd.DataFrame(rows, columns=columns)
    return df


