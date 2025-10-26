import utils
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

def create_sales_summary2(ws, head=1, start_row=2):
    # Ustun nomlarini aniqlaymiz
    headers = [cell.value for cell in ws[head]]
    date_idx = headers.index("Дата отчета")
    region_idx = headers.index("Регион")
    shop1_idx = headers.index("Продажи магазин 1")
    shop2_idx = headers.index("Продажи магазин 2")
    shop3_idx = headers.index("Продажи магазин 3")
    itogo_idx = headers.index("Итого")

    summary = defaultdict(lambda: {"count": 0, "sum": 0})

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        date_value = row[date_idx]
        if not date_value:
            continue
        if isinstance(date_value, str):
            try:
                date_value = datetime.strptime(date_value, "%d/%m/%Y")
            except:
                continue

        month = date_value.strftime("%B")  # Oy nomi
        region = row[region_idx]
        try:
            total_sales = sum([
                row[shop1_idx] or 0,
                row[shop2_idx] or 0,
                row[shop3_idx] or 0,
                row[itogo_idx] or 0
            ])
        except:
            continue
        key = (region, month)
        summary[key]["count"] += 1
        summary[key]["sum"] += total_sales

    # Natijani yangi sheetga yozamiz
    wb = ws.parent
    sheet_name = "Pivot - 8"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    pivot_ws = wb.create_sheet(sheet_name)

    pivot_ws.append(["Регион", "Месяц", "Количество", "Сумма продаж"])

    for (region, month), vals in summary.items():
        pivot_ws.append([region, month, vals["count"], vals["sum"]])

    return pivot_ws


def create_sales_summary(ws, head=1, start_row=2):
    # Ustun nomlarini aniqlaymiz
    headers = [cell.value for cell in ws[head]]
    date_idx = headers.index("Дата отчета")
    city_idx = headers.index("Город СБЕ")
    region_idx = headers.index("Регион")
    drug_idx = headers.index("Препарат")
    shop1_idx = headers.index("Продажи магазин 1")
    shop2_idx = headers.index("Продажи магазин 2")
    shop3_idx = headers.index("Продажи магазин 3")
    itogo_idx = headers.index("Итого")
    
    summary = defaultdict(lambda: {"count": 0, "sum": 0})

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        date_value = row[date_idx]
        if not date_value:
            continue
        if isinstance(date_value, str):
            try:
                date_value = datetime.strptime(date_value, "%d/%m/%Y")
            except:
                continue

        month = date_value.strftime("%B")  # Oy nomi
        city = row[city_idx]
        region = row[region_idx]
        drug = row[drug_idx]

        total_sales = sum([
            row[shop1_idx] or 0,
            row[shop2_idx] or 0,
            row[shop3_idx] or 0
        ])

        key = (city, region, drug, month)
        summary[key]["count"] += 1
        summary[key]["sum"] += total_sales

    # Natijani yangi sheetga yozamiz
    wb = ws.parent
    if "Pivot - 8" in wb.sheetnames:
        pivot_ws = wb["Pivot - 8"]
        wb.remove(pivot_ws)
    pivot_ws = wb.create_sheet("Pivot - 8")

    pivot_ws.append(["Город СБЕ", "Регион", "Препарат", "Месяц", "Количество", "Сумма продаж"])

    for (city, region, drug, month), vals in summary.items():
        pivot_ws.append([city, region, drug, month, vals["count"], vals["sum"]])

    return pivot_ws

# Panda ga import qilish
wb, ws = utils.excel.to_wb_and_ws(
    path="Works/Y2025/O10_K25/file.xlsx",
    sheet_name="Задание 8"
)
print(ws[1][1].value)

create_sales_summary(ws, 5, start_row=5)
# create_sales_summary2(ws, 5, start_row=5)

wb.save("Works/Y2025/O10_K25/output.xlsx")

